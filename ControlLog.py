from datetime import datetime
import re
from typing import List, Dict, Any
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

rootpath = 'D:/00_PROJECT/04_Python/LogAnalyzer/'  

motion_dict = {
    '528': 'ADIT_UCS',
    '258': 'SET_DO',
    '514': 'HOME_GRP',
    '515': 'STOP_GRP',
    '517': 'HALT',
    '518': 'CONTINUE',
    '519': 'ACK_GRP',
    '769': 'PUT_RDY_APP',
    '770': 'GET_RDY_APP',
    '771': 'TRG_RDY_APP',
    '772': 'EXTA_APP',
    '773': 'RETA_APP',
    '774': 'SAFE_R_APP',
    '775': 'SAFE_CST_APP',
    '776': 'TEACH_APP',
    '777': 'MPRDY_APP',
    '778': 'MPMOYION_APP',
    '780': 'ZPRDY_APP',
    '779': 'ZGRDY_APP',
    '781': 'RSWAP_APP',
    '782': 'NRDY_APP',
    '783': 'ROTATE_APP',
    '1025': 'PAR_UPDATE',
    '1026': 'CST_UPDATE',
    '1027': 'ADD_MONITOR',
    '1028': 'DEL_MONITOR',
    '523': 'JOG_JT',
    '527': 'JOG_STOP'
}

def parse_log_line(line: str) -> Dict[str, Any]:
    """解析单行日志"""
    try:
        # 正则表达式提取 CmdID、状态和动作
        pattern = r'(\d{2}:\d{2}:\d{2}\.\d{3})\s+\[CmdID / UniID = \[(\d+)\s+\]\[(\w+)\]\[(\d+)\s+(.*?)\]'
        match = re.match(pattern, line)
        if match:
            timestamp_str, cmd_id, status, action_code, action_stat = match.groups()
            return {
            'timestamp': datetime.strptime(timestamp_str, '%H:%M:%S.%f'),
            'cmd_id': cmd_id.strip(),
            'status': status.strip(),
            'action_code': action_code.strip(),
            'action_stat': action_stat.strip()
        }
        return None
        
    except Exception as e:
        print(f"Error parsing line: {line}")
        print(f"Error details: {str(e)}")
        return None

def parse_specific_actions(log_content: str) -> List[Dict[str, Any]]:
    """解析特定动作的开始和结束"""
    action_starts = {}
    completed_actions = []
    
    for line in log_content.split('\n'):
        if not line.strip():
            continue
            
        parsed = parse_log_line(line)
        if not parsed:
            continue
        
        cmd_id = parsed['cmd_id']
        action_stat = parsed['action_stat']
        status = parsed['status']
        timestamp = parsed['timestamp']
        
        # 处理开始事件
        if 'SHM_Updated' in action_stat:
            key = f"{cmd_id}_{timestamp.strftime('%H:%M:%S.%f')}"
            print(key)
            action_starts[key] = {
                'start_time': timestamp,
                'cmd_id': cmd_id,
                'action_stat': action_stat
            }
        
        # 处理结束事件
        elif 'Finish' in action_stat:
            # 查找匹配的开始事件
            matching_key = None
            for key in list(action_starts.keys()):
                if key.startswith(f"{cmd_id}_"):
                    matching_key = key
                    break
            
            if matching_key:
                start_info = action_starts[matching_key]
                duration = (timestamp - start_info['start_time']).total_seconds()
                
                completed_actions.append({
                    'action_stat': start_info['action_stat'],
                    'cmd_id': cmd_id,
                    'start_time': start_info['start_time'],
                    'end_time': timestamp,
                    'duration': duration
                })
                
                del action_starts[matching_key]
    
    # 按开始时间排序
    return sorted(completed_actions, key=lambda x: x['start_time'])

def create_excel_report(actions: List[Dict[str, Any]], output_file: str):
    """创建Excel报告"""
    # 准备时间线数据
    timeline_data = [{
        '开始时间': action['start_time'].strftime('%H:%M:%S.%f'),
        '结束时间': action['end_time'].strftime('%H:%M:%S.%f'),
        '指令ID': motion_dict[action['cmd_id']],
        '耗时(秒)': round(action['duration'], 3)
    } for action in actions]
    
    # 计算统计信息
    stats = {}
    for action in actions:
        name = action['cmd_id']
        if name not in stats:
            stats[name] = {
                'count': 0,
                'total_time': 0,
                'min_time': float('inf'),
                'max_time': 0
            }
        
        s = stats[name]
        s['count'] += 1
        s['total_time'] += action['duration']
        s['min_time'] = min(s['min_time'], action['duration'])
        s['max_time'] = max(s['max_time'], action['duration'])
    
    stats_data = [{
        '指令ID': name,
        '执行次数': s['count'],
        '平均耗时(秒)': round(s['total_time'] / s['count'], 3),
        '最短耗时(秒)': round(s['min_time'], 3),
        '最长耗时(秒)': round(s['max_time'], 3),
        '总耗时(秒)': round(s['total_time'], 3)
    } for name, s in stats.items()]
    
    # 创建Excel写入器
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入时间线sheet
        pd.DataFrame(timeline_data).to_excel(writer, sheet_name='动作时间线', index=False)
        
        # 写入统计信息sheet
        pd.DataFrame(stats_data).to_excel(writer, sheet_name='统计信息', index=False)
        
        # 格式化Excel
        workbook = writer.book
        
        for sheet_name in ['动作时间线', '统计信息']:
            sheet = workbook[sheet_name]
            
            # 设置列宽
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 设置标题行格式
            for cell in sheet[1]:
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

def main():
    try:
        # 读取日志文件
        with open(rootpath + 'ctrl.txt', 'r') as file:
            log_content = file.read()
        
        print("开始解析日志...")
        actions = parse_specific_actions(log_content)
        print(f"成功解析 {len(actions)} 个动作")
        
        print("\n创建Excel报告...")
        create_excel_report(actions, rootpath + '动作分析报告.xlsx')
        print("Excel报告已成功生成: 动作分析报告.xlsx")
        
        # 打印预览
        print("\n动作时间线预览 (前5条记录):")
        print("-" * 80)
        for action in actions[:5]:
            print(f"指令ID: {action['cmd_id']}")
            print(f"开始时间: {action['start_time'].strftime('%H:%M:%S.%f')}")
            print(f"结束时间: {action['end_time'].strftime('%H:%M:%S.%f')}")
            print(f"耗时: {action['duration']:.3f}秒")
            print("-" * 80)
    
    except Exception as e:
        print(f"程序执行出错: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    main()